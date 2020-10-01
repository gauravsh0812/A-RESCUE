import os
from openpyxl import Workbook

parameters = ['system.mem_ctrls.bytes_read::total', 'system.mem_ctrls.bytes_inst_read::total', 'system.mem_ctrls.num_reads::total', 'system.mem_ctrls.bw_read::total', 'system.mem_ctrls.bw_inst_read::total', 'system.mem_ctrls.bw_total::total', 'system.mem_ctrls.readReqs ','system.mem_ctrls.writeReqs','system.mem_ctrls.bytesReadDRAM','system.mem_ctrls.bytesWritten','system.mem_ctrls.totGap','system.mem_ctrls.totQLat', 'system.mem_ctrls.bytesPerActivate::total','system.mem_ctrls.totMemAccLat','system.mem_ctrls.totBusLat','system.mem_ctrls.avgQLat','system.mem_ctrls.avgBusLat','system.mem_ctrls.avgMemAccLat','system.mem_ctrls_0.totalEnergy', 'system.mem_ctrls_1.totalEnergy','system.cpu.dtb.flush_entries', 'system.cpu.itb.flush_entries', 'system.cpu.numCycles', 'system.cpu.num_int_register_reads','system.cpu.num_int_register_writes', 'system.cpu.num_fp_register_writes', 'system.cpu.num_fp_register_reads', 'system.cpu.num_cc_register_writes', 'system.cpu.num_cc_register_reads', 'system.cpu.dcache.tags.occ_percent::total','system.cpu.dcache.ReadReq_hits::total','system.cpu.dcache.WriteReq_hits::total','system.cpu.dcache.demand_hits::total','system.cpu.dcache.overall_hits::total','system.cpu.dcache.ReadReq_misses::total','system.cpu.dcache.WriteReq_misses::total','system.cpu.dcache.demand_misses::total','system.cpu.dcache.overall_misses::total','system.cpu.dcache.ReadReq_miss_latency::total','system.cpu.dcache.WriteReq_miss_latency::total','system.cpu.dcache.demand_miss_latency::total','system.cpu.dcache.overall_miss_latency::total','system.cpu.dcache.ReadReq_accesses::total','system.cpu.dcache.WriteReq_accesses::total','system.cpu.dcache.demand_accesses::total','system.cpu.dcache.overall_accesses::total','system.cpu.dcache.ReadReq_miss_rate::total','system.cpu.dcache.WriteReq_miss_rate::total','system.cpu.dcache.demand_miss_rate::total','system.cpu.dcache.overall_miss_rate::total','system.cpu.dcache.WriteReq_avg_miss_latency::total','system.cpu.dcache.demand_avg_miss_latency::total','system.cpu.dcache.overall_avg_miss_latency::total','system.cpu.dcache.ReadReq_mshr_misses::total','system.cpu.dcache.WriteReq_mshr_misses::total','system.cpu.dcache.demand_mshr_misses::total','system.cpu.dcache.overall_mshr_misses::total','system.cpu.dcache.ReadReq_mshr_miss_latency::total','system.cpu.dcache.WriteReq_mshr_miss_latency::total','system.cpu.dcache.demand_mshr_miss_latency::total','system.cpu.dcache.overall_mshr_miss_latency::total','system.cpu.dcache.ReadReq_mshr_miss_rate::total','system.cpu.dcache.WriteReq_mshr_miss_rate::total','system.cpu.dcache.demand_mshr_miss_rate::total','system.cpu.dcache.overall_mshr_miss_rate::total','system.cpu.dcache.ReadReq_avg_mshr_miss_latency::total','system.cpu.dcache.WriteReq_avg_mshr_miss_latency::total','system.cpu.dcache.demand_avg_mshr_miss_latency::total','system.cpu.dcache.overall_avg_mshr_miss_latency::total','system.cpu.icache.ReadReq_hits::total','system.cpu.icache.demand_hits::total','system.cpu.icache.overall_hits::total','system.cpu.icache.ReadReq_misses::total','system.cpu.icache.demand_misses::total','system.cpu.icache.overall_misses::total','system.cpu.icache.ReadReq_miss_latency::total','system.cpu.icache.demand_miss_latency::total','system.cpu.icache.overall_miss_latency::total','system.cpu.icache.ReadReq_accesses::total','system.cpu.icache.demand_accesses::total','system.cpu.icache.overall_accesses::total','system.cpu.icache.ReadReq_miss_rate::total','system.cpu.icache.demand_miss_rate::total','system.cpu.icache.overall_miss_rate::total','system.cpu.icache.ReadReq_avg_miss_latency::total','system.cpu.icache.demand_avg_miss_latency::total','system.cpu.icache.overall_avg_miss_latency::total','system.cpu.icache.ReadReq_mshr_misses::total','system.cpu.icache.demand_mshr_misses::total','system.cpu.icache.overall_mshr_misses::total','system.cpu.icache.ReadReq_mshr_miss_latency::total','system.cpu.icache.demand_mshr_miss_latency::total','system.cpu.icache.overall_mshr_miss_latency::total','system.cpu.icache.ReadReq_mshr_miss_rate::total','system.cpu.icache.demand_mshr_miss_rate::total','system.cpu.icache.overall_mshr_miss_rate::total','system.cpu.icache.ReadReq_avg_mshr_miss_latency::total','system.cpu.icache.demand_avg_mshr_miss_latency::total','system.cpu.icache.overall_avg_mshr_miss_latency::total','system.membus.pkt_count::total','system.membus.pkt_size::total','system.switch_cpus.numCycles','system.switch_cpus.committedInsts','system.switch_cpus.committedOps','system.switch_cpus.num_int_alu_accesses','system.switch_cpus.num_fp_alu_accesses','system.switch_cpus.num_func_calls','system.switch_cpus.num_int_register_reads','system.switch_cpus.num_int_register_writes','system.switch_cpus.num_fp_register_reads','system.switch_cpus.num_fp_register_writes','system.switch_cpus.num_cc_register_reads','system.switch_cpus.num_cc_register_writes','system.switch_cpus.op_class::total', 'system.cpu.dcache.overall_miss_rate::total', 'system.cpu.icache.overall_miss_rate::total']

para = []
value = []
for i in parameters:
    i=0
book = Workbook()
sheet = book.active
k=2
dir = (r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\d+i_cache\data_400M\SPEC_MI\stats_file')
os.chdir(dir)
#d_ret_time = ['10us', '26us', '50us', '75us', '100us', '1ms']#, '10ms', '100ms', '1s', '10s']
d_ret_time = ['400us']
i_ret_time = ['400us','1ms', '10ms', '50ms', '100ms', '1s']

app = ['astar', 'bwaves', 'bzip2', 'dijkstra', 'gamess', 'gemsfdtd', 'gobmk', 'gromacs', 'h264ref', 'hmmer', 'lbm', 'leslie3d', 'libquantum', 'mcf', 'milc', 'omnetpp', 'namd', 'povray', 'sjeng', 'soplex', 'tonto', 'xalancbmk', 'zeusmp', 'gsm', 'lame', 'm_cjpeg', 'm_djpeg', 'patricia']

for app in app:
    
    if app == 'gsm' or app == 'm_djpeg':
        phase = 1
    elif app== 'astar':
        phase = 3
    elif app == 'bwaves' or app == 'leslie3d':
        phase = 12
    elif app == 'bzip2':
        phase = 20
    elif app == 'gamess' or app == 'omnetpp':
        phase = 7
    elif app == 'gobmk' or app == 'soplex' or app == 'gemsfdtd':
        phase = 11
    elif app == 'gromacs' or app == 'h264ref' or app == 'mcf' or app == 'tonto' or app == 'm_cjpeg':
        phase = 8
    elif app == 'hmmer' or app == 'zeusmp':
        phase = 19
    elif app == 'lbm':
        phase = 16
    elif app == 'libquantum' or app == 'namd' or app == 'sjeng':
        phase = 5
    elif app == 'milc' or app == 'lame':
        phase = 15
    elif app == 'povray':
        phase = 4
    elif app == 'xalancbmk':
        phase == 13
    elif app == 'patricia':
        phase = 9
    elif app== 'dijkstra':
        phase = 2
    

    
    for t_d in d_ret_time:
        
        for t_i in i_ret_time:
            
            for i in range(1,phase+1):
                file = '{}_{}_{}_{}_64_32_4'.format(app,i,t_d, t_i) 
                for p in range(len(parameters)):
                    sheet.cell(row=1, column=p+5).value = parameters[p]
                                
                with open(file,'r') as fp:
                    for line in fp:
                        for j in range(len(parameters)):
                            
                            if parameters[j] in line:
                                end_index = line.find('#')
                                s = line[0:end_index-1]
                                s = s.strip()
                                s = s.split()
                                
                                sheet.cell(row=k, column=1).value = app
                                sheet.cell(row=k, column=2).value = t_d
                                sheet.cell(row=k, column=3).value = t_i
                                sheet.cell(row=k, column=4).value = i
                                
                                
                                sheet.cell(row=k, column=j+5).value = s[1]
                k+=1
       
                        
book.save('d+i_SPEC_MI.xlsx')
        
                    
                    
                                          
