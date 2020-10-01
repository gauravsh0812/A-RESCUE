# -*- coding: utf-8 -*-
"""
Created on Fri May 29 09:48:10 2020

@author: gaura
"""
import os
import pandas as pd 
import numpy as np 
import matplotlib.pyplot as plt
from sklearn.pipeline import Pipeline
import pickle
from sklearn.utils import shuffle
from sklearn.naive_bayes import GaussianNB
from sklearn.ensemble import AdaBoostClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.tree import DecisionTreeClassifier

for number in range(6, 10):

    df = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\d+i_cache\data_400M\d+i_COMBINE_DATA.xlsx', 'final_dataset')
    #df = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\d+i_cache\data_400M\MOBY\d+i_Moby.xlsx', 'Tosi_data_moby')
    
    df = df.fillna(df.mean()) 
    df = pd.DataFrame(df)
    
    # clustering algo 
    
    wcss = []
    from sklearn.cluster import KMeans
    for i in range(1, 15):
        kmeans = KMeans(n_clusters = i).fit(df.iloc[:, 4:-1])
        wcss.append(kmeans.inertia_)
    plt.plot(range(1, 15), wcss)
    #plt.show()
    
    
    kmeans = KMeans(n_clusters = 4, random_state = 0)
    y_kmeans = kmeans.fit_predict(df.iloc[:, 4:-1])
    
    cluster_0 = df.iloc[:, 4:-1][y_kmeans == 0].index.tolist()
    cluster_1 = df.iloc[:, 4:-1][y_kmeans == 1].index.tolist()
    cluster_2 = df.iloc[:, 4:-1][y_kmeans == 2].index.tolist()
    cluster_3 = df.iloc[:, 4:-1][y_kmeans == 3].index.tolist()
    
    clusters = [cluster_0,cluster_1,cluster_2,cluster_3]
    application_0, application_1, application_2, application_3 = [], [], [], []
    for n,i in enumerate(clusters):
        application = []
        for row in i:
            application.append(df.iloc[row, 0:4].values)
        if n ==0: application_0 = application
        if n ==1: application_1 = application
        if n ==2: application_2 = application
        if n ==3: application_3 = application        
    
    df = pd.concat([df[y_kmeans == 0]])
    feat_df = [i for i in df.columns if 'icache' in i]
    df_ind = [ind for ind, name in enumerate(df.columns.tolist()) if name in feat_df]
    df_y_e = df.loc[df.final_class == 'e'] 
    df_y_a = df.loc[df.final_class == 'a'] 
    df_y_b = df.loc[df.final_class == 'b']
    df_y_c = df.loc[df.final_class == 'c']
    df_y_d = df.loc[df.final_class == 'd']
    
    df_y_e = shuffle(df_y_e)
    df_y_e = df_y_e.iloc[0:20, :]
    df = pd.concat([df_y_a, df_y_b, df_y_c, df_y_d, df_y_e], axis = 0)
    df = shuffle(df)
    X = df.iloc[:, 4:-1]
    y = df.iloc[:, -1]
    from sklearn.feature_selection import SelectKBest, chi2
    
    best_features = SelectKBest(score_func = chi2, k = 30)
    fit = best_features.fit(X,y)
    
    # finding the scores and columns to concatenate them into a table
    
    scores = pd.DataFrame(fit.scores_)  
    columns = pd.DataFrame(pd.DataFrame(X).columns)
    
    score_table = pd.concat([columns, scores], axis = 1, ignore_index = True)
    score_table.columns = ['specs', 'score']
    sc_table = score_table.nlargest(25, 'score')
    index_sc_table = list(pd.DataFrame(sc_table).index.values)
    from sklearn.preprocessing import StandardScaler
    sc_x = StandardScaler()
    X = sc_x.fit_transform(X)
    from sklearn.model_selection import train_test_split
    x_train, x_test, y_train, y_test = train_test_split(X, y, test_size = 0.3, random_state = 0)
    
    x_train = pd.DataFrame(x_train)
    x_test = pd.DataFrame(x_test)
    
    ind= [13,31, 32, 11, 29,66, 67, 68,  5, 48, 50, 14, 10, 18, 49]
    
    
    X_train = x_train.iloc[:,ind]
    X_test = x_test.iloc[:,ind]
    
    pipeline_feat_DT= Pipeline([('classifier_feat_DT', DecisionTreeClassifier( criterion = 'entropy', random_state = 0))])
    pipeline_feat_RF= Pipeline([('classifier_feat_RF', RandomForestClassifier(n_estimators = 300, criterion = 'entropy', random_state = 0))])
    pipeline_feat_LT = Pipeline([('classifier_feat_LT', LogisticRegression(random_state = 0))])
    pipeline_feat_SVC = Pipeline([('classifier_feat_SVC', SVC(C = 1, kernel = 'rbf', gamma =0.1 , random_state = 0, cache_size = 200))])
    pipeline_feat_Knn = Pipeline([('classifier_feat_Knn', KNeighborsClassifier(n_neighbors = 3, metric = 'minkowski', p = 2))])
    pipeline_feat_Naive_Bayes = Pipeline([('classifier_feat_Naive_Bayes', GaussianNB())])
    pipeline_feat_Adaboost = Pipeline([('classifier_feat_Aaboost', AdaBoostClassifier(n_estimators=300, learning_rate=1).fit(X_train, y_train))])
    
    
    pipe_feat_list = [pipeline_feat_LT, pipeline_feat_DT, pipeline_feat_RF, pipeline_feat_SVC, pipeline_feat_Knn, pipeline_feat_Naive_Bayes, pipeline_feat_Adaboost]
    
    pipe_feat_dict = {0: 'LogisticRegression_Feat', 1: 'DecisionTree_Feat', 2: 'RandomForest_Feat', 3: 'SVC_Feat', 4: 'K-NN_Feat', 5: 'Naive_Bayes_Feat', 6:'Adaboost'}
    
    
        
    for pipe in pipe_feat_list:
        pipe.fit(X_train, y_train)
        
    
    Feat_table = []
    for i, model in enumerate(pipe_feat_list):
        accuracy = '{}_accuracy {}'.format(pipe_feat_dict[i], model.score(X_test, y_test))
        Feat_table.append(accuracy.split(' ')[1])
        
    col_1 = ['LogisticRegression', 'Decision Tree', 'Random Forest', 'SVC', 'K-NN', 'Naive_Bayes', 'Adaboost' ]
    col_2 = Feat_table
    Final_table = pd.concat([pd.DataFrame(col_1), pd.DataFrame(col_2)], axis = 1, ignore_index = True)
    Final_table.columns = ['ML_Classifier' , 'Feat_accuracy'] 
    Final_table = pd.DataFrame(Final_table)
    print(Final_table)
    
    #knn tuning 
    clf_knn = KNeighborsClassifier(n_neighbors = 3, metric = 'minkowski', p = 1).fit(X_train, y_train)
    print('Before tuning Knn: {}'.format(clf_knn.score(X_test, y_test)))
    
    from sklearn.model_selection import GridSearchCV
    
    parameters = [
                  {'n_neighbors' : [1, 2, 3, 4, 5, 10], 'metric':['minkowski'], 'p':[1, 2]},
                  
                    ]
    GS_knn = GridSearchCV(estimator = clf_knn, param_grid = parameters, scoring = 'accuracy', cv = 10, n_jobs = -1 )
    GS_knn.fit(X_train, y_train)
    
    accuracy_knn = GS_knn.best_score_
    print("Knn GS_cv accuracy: {}".format(accuracy_knn))
    print("GS_cv_knn best params: {}".format(GS_knn.best_params_))
    
    # Rf tuning 
    
    clf_RF= RandomForestClassifier(n_estimators = 20, criterion = 'entropy', random_state = 0).fit(X_train, y_train)
    clf_RF.score(X_test, y_test)
    
    parameters = [
                  {'n_estimators': [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['gini'],'max_features' :['auto'], 'random_state' : [0] }, 
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion': ['gini'], 'max_features' : ['sqrt'], 'random_state' : [0]},
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['gini'], 'max_features' : ['log2'], 'random_state' : [0]},
                  {'n_estimators': [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['entropy'],'max_features' :['auto'], 'random_state' : [0] }, 
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion': ['entropy'], 'max_features' : ['sqrt'],  'random_state' : [0]},
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['entropy'], 'max_features' : ['log2'],  'random_state' : [0]}]
    
    GS_RF = GridSearchCV(estimator = clf_RF, param_grid = parameters, scoring = 'accuracy', cv = 10, n_jobs = -1 )
    GS_RF.fit(X_train, y_train)
    
    accuracy_RF = GS_RF.best_score_
    print("RF GS_cv accuracy: {}".format(accuracy_RF))
    print("GS_cv_RF best params: {}".format(GS_RF.best_params_))
    
    # applying the parameters got from the tuning
    
    clf_RF= RandomForestClassifier(n_estimators = GS_RF.best_params_['n_estimators'], max_features = GS_RF.best_params_['max_features'], criterion = GS_RF.best_params_['criterion'], random_state = 0).fit(X_train, y_train)
    clf_RF.score(X_test, y_test)
    print('RF_tuned accuracy: {}'.format(clf_RF.score(X_test, y_test)))
    with open('i_RF_Tuned.sav', 'wb') as iDTt:
        pickle.dump(clf_RF, iDTt)
    
    clf_knn = KNeighborsClassifier(n_neighbors = GS_knn.best_params_['n_neighbors'], metric = GS_knn.best_params_['metric'], p = GS_knn.best_params_['p']).fit(X_train, y_train)
    print('knn_tuned accuracy: {}'.format(clf_knn.score(X_test, y_test)))
    with open('i_knn_Tuned.sav', 'wb') as iDTt:
        pickle.dump(clf_knn, iDTt)
    
    clf_DT_tuned= DecisionTreeClassifier(criterion = 'entropy', random_state = 0).fit(X_train, y_train)
    print('DT_tuned_accuracy: {}'.format(clf_DT_tuned.score(X_test, y_test)))
    
    
    #with open('i_DT_Tuned.sav', 'wb') as iDTt:
    #    pickle.dump(clf_DT_tuned, iDTt)
    
    os.mkdir(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\RF_DT\icache\Round_{}'.format(number))
    
    Final_table.loc[Final_table['ML_Classifier'] == 'Decision Tree', 'Feat_accuracy'] = clf_DT_tuned.score(X_test, y_test)
    Final_table.loc[Final_table['ML_Classifier'] == 'Random Forest', 'Feat_accuracy'] = clf_RF.score(X_test, y_test)
    Final_table.loc[Final_table['ML_Classifier'] == 'K-NN', 'Feat_accuracy'] = clf_knn.score(X_test, y_test)
    
    Final_table.to_csv(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\RF_DT\icache\accuracy_table_{}.csv'.format(number), index = False)
    
    y_pred_RF = clf_RF.predict(X_test)
    y_pred_array_RF = np.array(y_pred_RF)
    y_pred_RF
    y_pred_DT = clf_DT_tuned.predict(X_test)
    y_pred_array_DT = np.array(y_pred_DT)
    y_pred_DT
    y_pred_knn = clf_knn.predict(X_test)
    y_pred_array_knn = np.array(y_pred_knn)
    y_pred_knn
    y_test_array = np.array(y_test)
    y_test_array
    
    from openpyxl import Workbook
    
    df_alldata = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\d+i_cache\data_400M\d+i_COMBINE_DATA.xlsx', 'LE_calc')
    pred_arr = [y_pred_array_DT, y_pred_array_RF, y_pred_array_knn]
    name = ['DT', 'RF', 'knn']
    
    for array, name in zip(pred_arr, name):
    
        index = y_test.index
        df_array = []
        df_phase = []
        df_names = df_alldata.iloc[:, 0]
        df_phases = df_alldata.iloc[:, 3]
        for i in index:
            df_array.append(df_names[index])
            df_phase.append(df_phases[index])
        
        
        book = Workbook()
        sheet= book.active
        
        df_array = pd.DataFrame(df_array)
        df_phase = pd.DataFrame(df_phase)
        f = df_array.T.iloc[:,0]
        f_phs = df_phase.T.iloc[:,0]
        f.columns = ['index', 'app_name']
        f_phs.columns = ['index','phase']
        
        moby_arr = ['360buy', 'adobe', 'baidumap', 'bbench', 'frozenbubble', 'k9mail', 'netease', 'sinaweibo', 'mxplayer', 'ttpod' , 'kingsoftoffice']
    
        k = 2
        
        sheet.cell(row = 1, column = 1).value = 'application'
        sheet.cell(row = 1, column = 2).value = 'phase'
        sheet.cell(row = 1, column = 3).value = 'Predicted_class'
        sheet.cell(row = 1, column = 4).value = 'Actual_class'
        sheet.cell(row = 1, column = 5).value = 'i_PREDICTED_LAT_RET'
        sheet.cell(row = 1, column = 6).value = 'i_PRED_LAT_RET'
        sheet.cell(row = 1, column = 7).value = 'i_ACTUAL_ENERGY_RET'
        sheet.cell(row = 1, column = 8).value = 'i_ACTUAL_ENERGY_RET'
        sheet.cell(row = 1, column = 9).value = 'i_Latency_predicted'
        sheet.cell(row = 1, column = 10).value = 'i_Latency_actual'
        sheet.cell(row = 1, column = 11).value = 'i_Energy_predicted'
        sheet.cell(row = 1, column = 12).value = 'i_Energy_actual'
        
        for index, i, j in zip(index, array, y_test_array):
            
            if f[index] in moby_arr:
                
                sheet['A{}'.format(k)] = f[index]
                sheet['B{}'.format(k)] = f_phs[index]
                sheet['C{}'.format(k)] = i
                sheet['D{}'.format(k)] = j
                
                
                if i == 'a':
                    pair = ['1ms']
                    
                elif i == 'b':
                    pair = ['50ms']
                
                elif i == 'c':
                    pair = ['100ms']
                
                elif i == 'd':
                    pair = ['1s']
                
                elif i == 'e':
                    pair = ['10ms']
                
               
                if j == 'a':
                    pair_2 = ['1ms']
                    
                elif j == 'b':
                    pair_2 = ['50ms']
                
                elif j == 'c':
                    pair_2 = ['100ms']    
                
                elif j == 'd':
                    pair_2 = ['1s']
                
                elif i == 'e':
                    pair_2 = ['10ms']
                    
                lat_ret_pred = pair[0]
                energy_ret_pred = pair[0]    
                lat_ret_actual = pair_2[0]
                energy_ret_actual = pair_2[0]   
                
                
                df_1 = df_alldata
                df_1 = df_1.loc[:,['application','phase no.','d_ret_time', 'i_ret_time', 'i_latency','i_energy']]
                
                df_row_lat_pred = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['i_ret_time'] == lat_ret_pred)]
                df_row_energy_pred = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['i_ret_time'] == energy_ret_pred)]
                df_row_lat_actual = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['i_ret_time'] == lat_ret_actual )]
                df_row_energy_actual = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['i_ret_time'] == energy_ret_actual)]
                
                # for base = 1s
                
                df_row_lat_1s = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['i_ret_time'] == '1s')]
                df_row_energy_1s = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['i_ret_time'] == '1s')]
                
                #exhautive
                df_row_lat_exh = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index])]
                df_row_energy_exh = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index])]
            
            
                i_Lat_pred = df_row_lat_pred.iloc[:,-2]
                i_Lat_pred = np.array(i_Lat_pred)
                i_Lat_pred = i_Lat_pred[0]
                
                i_energy_pred = df_row_energy_pred.iloc[:,-1]
                i_energy_pred = np.array(i_energy_pred)
                i_energy_pred = i_energy_pred[0]
                
                i_Lat_actual = df_row_lat_actual.iloc[:,-2]
                i_Lat_actual = np.array(i_Lat_actual)
                i_Lat_actual = i_Lat_actual[0]
                
                i_energy_actual = df_row_energy_actual.iloc[:,-1]
                i_energy_actual = np.array(i_energy_actual)
                i_energy_actual = i_energy_actual[0]
                
                i_exh_lat = min(df_row_lat_exh.iloc[:,-2].values)
                i_exh_energy = min(df_row_energy_exh.iloc[:,-1].values)
                
                i_Lat_1s = df_row_lat_1s.iloc[:,-2]
                i_Lat_1s = np.array(i_Lat_1s)
                i_Lat_1s = i_Lat_1s[0]
                
                i_energy_1s = df_row_energy_1s.iloc[:,-1]
                i_energy_1s = np.array(i_energy_1s)
                i_energy_1s = i_energy_1s[0]
                
                sheet['E{}'.format(k)] = lat_ret_pred
                sheet['F{}'.format(k)] = energy_ret_pred
                sheet['I{}'.format(k)] = i_Lat_pred
                sheet['J{}'.format(k)] = i_Lat_actual
                sheet['K{}'.format(k)] = i_energy_pred
                sheet['L{}'.format(k)] = i_energy_actual
                sheet['G{}'.format(k)] = lat_ret_actual
                sheet['H{}'.format(k)] = energy_ret_actual
                sheet['M{}'.format(k)] = i_exh_lat
                sheet['N{}'.format(k)] = i_exh_energy
                sheet['O{}'.format(k)] = i_Lat_1s
                sheet['P{}'.format(k)] = i_energy_1s
                sheet['R{}'.format(k)] = (i_Lat_1s - i_Lat_pred)/i_Lat_1s
                sheet['T{}'.format(k)] = (i_Lat_1s - i_Lat_actual)/i_Lat_1s
                sheet['V{}'.format(k)] = (i_Lat_1s - i_exh_lat)/i_Lat_1s
                sheet['S{}'.format(k)] = (i_energy_1s - i_energy_pred)/i_energy_1s
                sheet['U{}'.format(k)] = (i_energy_1s - i_energy_actual)/i_energy_1s
                sheet['W{}'.format(k)] = (i_energy_1s - i_exh_energy)/i_energy_1s 
                k+=1
                print(name, f[index],i, j, i_Lat_pred, i_Lat_actual)
        
        book.save(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\RF_DT\icache\Round_{}\icache_{}_{} .xlsx'.format(number, name, number))
