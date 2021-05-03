
import pandas as pd 
import numpy as np
import matplotlib.pyplot as plt
import pickle
import os
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import cross_val_score
from sklearn.utils import shuffle

for number in range(12, 21):
    
    # Importing the dataset
    
    df = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\old\DATA_Classifier.xlsx', 'data_75us')
    df = shuffle(df)
    X = df.iloc[:, 3:119]
    y = df.iloc[:, -1]
    X = pd.DataFrame(X).fillna(0)
    
    from sklearn.feature_selection import SelectKBest, chi2
    
    best_features = SelectKBest(score_func = chi2, k = 30)
    fit = best_features.fit(X,y)
    
    # finding the scores and columns to concatenate them into a table
    
    scores = pd.DataFrame(fit.scores_)  
    columns = pd.DataFrame(X.columns)
    
    score_table = pd.concat([columns, scores], axis = 1, ignore_index = True)
    score_table.columns = ['specs', 'score']
    sc_table = score_table.nlargest(25, 'score')
    index_sc_table = list(pd.DataFrame(sc_table).index.values)
    
    col = [39, 58, 60,75,76,77,90,91,92,5,4,19,10, 13]
    
    from sklearn.preprocessing import StandardScaler
    sc_x = StandardScaler()
    X = sc_x.fit_transform(X)
    from sklearn.model_selection import train_test_split
    x_train, x_test, y_train, y_test = train_test_split(X, y, test_size = 0.3, random_state = 0)
    
    '''
    from sklearn.decomposition import PCA
    pca = PCA().fit(X)
    
    # plotting of Cumulative Summation of explained variance ratio 
    
    plt.figure()
    plt.plot(np.cumsum(pca.explained_variance_ratio_))
    plt.xlim(-1,30)
    plt.xlabel('no. of PC')
    plt.ylabel('variance %age')
    plt.title('Dataset Explained Variance Ratio')
    plt.show()
    
    pipeline_pca_LT = Pipeline([('pca_LT', PCA(n_components = 20)),
                               ('classifier_pca_LT', LogisticRegression(random_state = 0))])
    pipeline_pca_DT = Pipeline([('pca_DT', PCA(n_components = 20)),
                               ('classifier_pca_DT', DecisionTreeClassifier( criterion = 'entropy', random_state = 0))])
    pipeline_pca_RF = Pipeline([('pca_RF', PCA(n_components = 20)),
                               ('classifier_pca_RF', RandomForestClassifier(n_estimators = 50, criterion = 'entropy', random_state = 0))])
    pipeline_pca_SVC = Pipeline([('pca_SVC', PCA(n_components = 20)),
                               ('classifier_pca_SVC', SVC(C = 1, kernel = 'rbf', gamma =0.1 , random_state = 0, cache_size = 200))])
    pipeline_pca_Knn = Pipeline([('pca_Knn', PCA(n_components = 20)),
                               ('classifier_pca_Knn', KNeighborsClassifier(n_neighbors = 5, metric = 'minkowski', p = 2))])
    pipeline_pca_Naive_Bayes = Pipeline([('pca_Naive_Bayes', PCA(n_components = 20)),
                               ('classifier_pca_Naive_Bayes', GaussianNB())])
    '''
    x_train = pd.DataFrame(x_train)
    x_test = pd.DataFrame(x_test)
    
    x_feat_train = x_train.iloc[:,col]
    x_feat_test = x_test.iloc[:,col]
    
    pipeline_feat_DT= Pipeline([('classifier_feat_DT', DecisionTreeClassifier( criterion = 'entropy', random_state = 0))])
    pipeline_feat_RF= Pipeline([('classifier_feat_RF', RandomForestClassifier(n_estimators = 300, criterion = 'entropy', random_state = 0))])
    pipeline_feat_LT = Pipeline([('classifier_feat_LT', LogisticRegression(random_state = 0))])
    pipeline_feat_SVC = Pipeline([('classifier_feat_SVC', SVC(C = 1, kernel = 'rbf', gamma =0.1 , random_state = 0, cache_size = 200))])
    pipeline_feat_Knn = Pipeline([('classifier_feat_Knn', KNeighborsClassifier(n_neighbors = 5, metric = 'minkowski', p = 2))])
    pipeline_feat_Naive_Bayes = Pipeline([('classifier_feat_Naive_Bayes', GaussianNB())])
    
    #pipe_pca_list = [pipeline_pca_LT, pipeline_pca_DT, pipeline_pca_RF, pipeline_pca_SVC, pipeline_pca_Knn, pipeline_pca_Naive_Bayes ]
    
    pipe_feat_list = [pipeline_feat_LT, pipeline_feat_DT, pipeline_feat_RF, pipeline_feat_SVC, pipeline_feat_Knn, pipeline_feat_Naive_Bayes]
    
    #pipe_pca_dict = {0: 'LogisticRegression_PCA', 1: 'DecisionTree_PCA', 2: 'RandomForest_PCA', 3: 'SVC_PCA', 4: 'K-NN_PCA', 5: 'Naive_Bayes_PCA'}
    pipe_feat_dict = {0: 'LogisticRegression_Feat', 1: 'DecisionTree_Feat', 2: 'RandomForest_Feat', 3: 'SVC_Feat', 4: 'K-NN_Feat', 5: 'Naive_Bayes_Feat'}
    
    #for pipe in pipe_pca_list:
     #   pipe.fit(x_train, y_train)
    
        
    for pipe in pipe_feat_list:
        pipe.fit(x_feat_train, y_train)
        
    '''
    PCA_table = []
    for i, model in enumerate(pipe_pca_list):
        
        accuracy = '{}_accuracy {}'.format(pipe_pca_dict[i], model.score(x_test, y_test))
        PCA_table.append(accuracy.split(' ')[1])
        #print(accuracy)
    ''' 
    Feat_table = []
    for i, model in enumerate(pipe_feat_list):
        accuracy = '{}_accuracy {}'.format(pipe_feat_dict[i], model.score(x_feat_test, y_test))
        Feat_table.append(accuracy.split(' ')[1])
        
    col_1 = ['LogisticRegression', 'Decision Tree', 'Random Forest', 'SVC', 'K-NN', 'Naive_Bayes' ]
    #col_2 = PCA_table
    col_3 = Feat_table
    Final_table = pd.concat([pd.DataFrame(col_1), pd.DataFrame(col_3)], axis = 1, ignore_index = True)
    Final_table.columns = ['ML_Classifier' ,'Feat_accuracy'] 
    print(Final_table)
    adaboost = []
    from sklearn.ensemble import AdaBoostClassifier
    clf_ada = AdaBoostClassifier(n_estimators=300, learning_rate=1).fit(x_feat_train, y_train)
    print('accuracy for Adaboost: {}'.format(clf_ada.score(x_feat_test, y_test)))
    adaboost.append(clf_ada.score(x_feat_test, y_test))
    #with open('adaboost.pickle', 'wb') as RF:
     #       pickle.dump(clf_ada, RF)
     
     
    # tuning RF
     
    RF_Classifier = RandomForestClassifier(n_estimators = 300, criterion = 'entropy', random_state = 0)
    RF_Classifier.fit(x_feat_train, y_train)
    y_pred_RFC = RF_Classifier.predict(x_feat_test)
    from sklearn.metrics import confusion_matrix, accuracy_score
    cm_RFC = confusion_matrix(y_test, y_pred_RFC)
    accuracy_RF = accuracy_score(y_test, y_pred_RFC)
    print('accuracy RF befor tuning: {}'.format(accuracy_RF))
    
    from sklearn.model_selection import GridSearchCV
    
    parameters = [
                  {'n_estimators': [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['gini'],'max_features' :['auto'], 'random_state' : [0] }, 
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion': ['gini'], 'max_features' : ['sqrt'], 'random_state' : [0]},
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['gini'], 'max_features' : ['log2'], 'random_state' : [0]},
                  {'n_estimators': [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['entropy'],'max_features' :['auto'], 'random_state' : [0] }, 
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion': ['entropy'], 'max_features' : ['sqrt'],  'random_state' : [0]},
                  {'n_estimators' : [20, 30, 50, 70, 100, 200, 300, 500, 1000], 'criterion' : ['entropy'], 'max_features' : ['log2'],  'random_state' : [0]}]
    
    GS_RF = GridSearchCV(estimator = RF_Classifier, param_grid = parameters, scoring = 'accuracy', cv = 10, n_jobs = -1 )
    GS_RF.fit(x_feat_train, y_train)
    accuracy = GS_RF.best_score_
    print('accuracy_GS_RF: {}'.format(accuracy))
    print('best params for RF: {}'.format(GS_RF.best_params_))
    RF_Classifier_tuned = RandomForestClassifier(n_estimators = GS_RF.best_params_['n_estimators'], max_features = GS_RF.best_params_['max_features'], criterion = GS_RF.best_params_['criterion'], random_state = 0).fit(x_feat_train, y_train)
    RF_Classifier_tuned.fit(x_feat_train, y_train)
    
    y_pred_tuned = RF_Classifier_tuned.predict(x_feat_test)
    cm_tuned = confusion_matrix(y_test, y_pred_tuned)
    accuracy_RF_tuned = accuracy_score(y_test, y_pred_tuned)
    print('accuracy_RF_tuned:{}'.format(accuracy_RF_tuned))
    
    # DT tuning
    
    parameters = [
                  {'criterion' : ['gini'],'splitter' :['best'], 'random_state' : [0] }, 
                  {'criterion' : ['gini'], 'splitter' : ['random'], 'random_state' : [0]},
                  {'criterion' : ['entropy'],'splitter' :['best'], 'random_state' : [0] }, 
                  {'criterion' : ['entropy'], 'splitter' : ['random'], 'random_state' : [0]},]
    
    GS_RF = GridSearchCV(estimator = RF_Classifier, param_grid = parameters, scoring = 'accuracy', cv = 10, n_jobs = -1 )
    GS_RF.fit(x_feat_train, y_train)
    accuracy = GS_RF.best_score_
    print('accuracy_GS_RF: {}'.format(accuracy))
    print('best params for RF: {}'.format(GS_RF.best_params_))
    RF_Classifier_tuned = RandomForestClassifier(n_estimators = GS_RF.best_params_['n_estimators'], max_features = GS_RF.best_params_['max_features'], criterion = GS_RF.best_params_['criterion'], random_state = 0).fit(x_feat_train, y_train)
    RF_Classifier_tuned.fit(x_feat_train, y_train)
    
    y_pred_tuned = RF_Classifier_tuned.predict(x_feat_test)
    cm_tuned = confusion_matrix(y_test, y_pred_tuned)
    accuracy_RF_tuned = accuracy_score(y_test, y_pred_tuned)
    print('accuracy_RF_tuned:{}'.format(accuracy_RF_tuned))
    
    from sklearn.tree import DecisionTreeClassifier
    DT = DecisionTreeClassifier( criterion = 'entropy', random_state = 0)
    DT.fit(x_feat_train, y_train)
    
    y_pred_DT = DT.predict(x_feat_test)
    
    from sklearn.metrics import confusion_matrix, accuracy_score
    cm_DT = confusion_matrix(y_test, y_pred_DT)
    accuracy_DT_tuned = accuracy_score(y_test, y_pred_DT)
    print('accuracy_DT_tuned: {}'.format(accuracy_DT_tuned))
    
    '''
    # Knn tuning 
    
    GS_knn = KNeighborsClassifier(n_neighbors = 5, metric = 'minkowski', p = 2)
    GS_knn.fit(x_feat_train, y_train)
    parameters = [
                  {'n_neighbors' : [1, 2, 3, 4, 5, 10], 'metric':['minkowski'], 'p':[1, 2]},
                  
                    ]
    GS_knn = GridSearchCV(estimator = GS_knn, param_grid = parameters, scoring = 'accuracy', cv = 10, n_jobs = -1 )
    GS_knn.fit(x_feat_train, y_train)
    
    accuracy_knn_GS = GS_knn.best_score_
    print("Knn GS_cv accuracy: {}".format(accuracy_knn_GS))
    print("GS_cv_knn best params: {}".format(GS_knn.best_params_))
    
    y_pred_knn = GS_knn.predict(x_feat_test)
    cm_knn = confusion_matrix(y_test, y_pred_knn)
    accuracy_knn_tuned = accuracy_score(y_test, y_pred_knn)
    print('accuracy tuned_knn: {}'.format(accuracy_knn_tuned))
    '''
    os.mkdir(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\RF_DT\dcache\Round_{}'.format(number))
    
    Final_table.loc[Final_table['ML_Classifier'] == 'Decision Tree', 'Feat_accuracy'] = accuracy_DT_tuned
    Final_table.loc[Final_table['ML_Classifier'] == 'Random Forest','Feat_accuracy'] = accuracy_RF_tuned
    Final_table.loc[Final_table['ML_Classifier'] == 'K-NN','Feat_accuracy'] = accuracy_knn_tuned
    
    Final_table.to_csv(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\RF_DT\dcache\accuracy_table_{}.csv'.format(number), index = False)
    
    
    #excel sheet
    
    y_pred_array_RF = np.array(y_pred_tuned)
    y_pred_array_DT = np.array(y_pred_DT)
    y_pred_array_knn = np.array(y_pred_knn)
    y_test_array = np.array(y_test)
    df_alldata = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\old\DATA_FOR_CLASSIFIER.xlsx', 'data')
    
    
    
    
    pred_arr = [y_pred_array_DT]#, y_pred_array_RF]
    name = ['DT']#, 'RF']
    
    for array, name in zip(pred_arr, name):
        
        index = y_test.index
        df_array = []
        df_phase = []
        df_names = df.iloc[:, 0]
        df_phases = df.iloc[:, 2]
        for i in index:
            df_array.append(df_names[index])
            df_phase.append(df_phases[index])
        
        from openpyxl import Workbook
        
        book = Workbook()
        sheet= book.active
        df_array = pd.DataFrame(df_array)
        df_phase = pd.DataFrame(df_phase)
        f = df_array.T.iloc[:,0]
        f_phs = df_phase.T.iloc[:,0]
        f.columns = ['index', 'app_name']
        f_phs.columns = ['index','phase']
        
    
        k = 2
        
        sheet.cell(row = 1, column = 1).value = 'application'
        sheet.cell(row = 1, column = 2).value = 'phase'
        sheet.cell(row = 1, column = 3).value = 'Predicted_class'
        sheet.cell(row = 1, column = 4).value = 'Actual_class'
        sheet.cell(row = 1, column = 5).value = 'PREDICTED_LAT_RET'
        sheet.cell(row = 1, column = 6).value = 'ACTUAL_LAT_RET'
        sheet.cell(row = 1, column = 7).value = 'PRED_ENERGY_RET'
        sheet.cell(row = 1, column = 8).value = 'ACTUAL_ENERGY_RET'
        sheet.cell(row = 1, column = 9).value = 'Latency_predicted'
        sheet.cell(row = 1, column = 10).value = 'Latency_actual'
        sheet.cell(row = 1, column = 11).value = 'Energy_predicted'
        sheet.cell(row = 1, column = 12).value = 'Energy_actual'
        sheet.cell(row = 1, column = 13).value = 'exh_lat'
        sheet.cell(row = 1, column = 14).value = 'exh_energy'
        sheet.cell(row = 1, column = 15).value = '1ms_lat'
        sheet.cell(row = 1, column = 16).value = '1ms_energy'
        
        df_1 = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\old\DATA_Classifier.xlsx','lat_energy_calc')  
        df_1 = df_1.loc[:,['application','phase no.','ret_time','Latency','ENERGY']]
            
        for index, i, j in zip(index, y_pred_array_RF, y_test_array):
            
            sheet['A{}'.format(k)] = f[index]
            sheet['B{}'.format(k)] = f_phs[index]
            sheet['C{}'.format(k)] = i
            sheet['D{}'.format(k)] = j
            
            
            if i == 'D':
                pair = ['10us', '26us']
                
            elif i == 'E':
                pair = ['75us', '1ms']
            
            elif i == 'F':
                pair = ['50us', '75us']
            
            elif i == 'G':
                pair = ['1ms', '75us', '50us']
            
            elif i == 'H':
                pair = ['10us','26ms', '50us']
            
            elif i == 'I':
                pair = ['26us', '50ms', '75us']
                
            elif i == 'J':
                pair = ['10us', '26us', '50us', '75us']    
            
            if j == 'D':
                pair_2 = ['10us', '26us']
                
            elif j == 'E':
                pair_2 = ['75us', '1ms']
            
            elif j == 'F':
                pair_2 = ['50us', '75us']
            
            elif j == 'G':
                pair_2 = ['1ms', '75us', '50us']
            
            elif j == 'H':
                pair_2 = ['10us','26ms', '50us']
            
            elif j == 'I':
                pair_2 = ['26us', '50ms', '75us']
                
            elif j == 'J':
                pair_2 = ['10us', '26us', '50us', '75us']    
            
            lat_ret_pred = pair[-1]
            energy_ret_pred = pair[0]    
            lat_ret_actual = pair_2[-1]
            energy_ret_actual = pair_2[0]   
            
        
            df_row_lat_pred = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == lat_ret_pred)]
            df_row_energy_pred = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == energy_ret_pred)]
            df_row_lat_actual = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == lat_ret_actual )]
            df_row_energy_actual = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == energy_ret_actual)]
            
            # for base = 1ms
            
            df_row_lat_1ms = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == '1ms')]
            df_row_energy_1ms = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == '1ms')]
            
            #exhautive
            df_row_lat_exh = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index])]
            df_row_energy_exh = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index])]
            
            
            Lat_pred = df_row_lat_pred.iloc[:,3]
            Lat_pred = np.array(Lat_pred)
            Lat_pred = Lat_pred[0]
            
            energy_pred = df_row_energy_pred.iloc[:,4]
            energy_pred = np.array(energy_pred)
            energy_pred = energy_pred[0]
            
            Lat_actual = df_row_lat_actual.iloc[:,3]
            Lat_actual = np.array(Lat_actual)
            Lat_actual = Lat_actual[0]
            
            energy_actual = df_row_energy_actual.iloc[:,4]
            energy_actual = np.array(energy_actual)
            energy_actual = energy_actual[0]
            
            Lat_1ms = df_row_lat_1ms.iloc[:,3]
            Lat_1ms = np.array(Lat_1ms)
            Lat_1ms = Lat_1ms[0]
            
            energy_1ms = df_row_energy_1ms.iloc[:,4]
            energy_1ms = np.array(energy_1ms)
            energy_1ms = energy_1ms[0]
            
            exh_lat = min(df_row_lat_exh.iloc[:,-2].values)
            exh_energy = min(df_row_energy_exh.iloc[:,-1].values)
            
            sheet['E{}'.format(k)] = lat_ret_pred
            sheet['F{}'.format(k)] = energy_ret_pred
            sheet['I{}'.format(k)] = Lat_pred
            sheet['J{}'.format(k)] = Lat_actual
            sheet['K{}'.format(k)] = energy_pred
            sheet['L{}'.format(k)] = energy_actual
            sheet['G{}'.format(k)] = lat_ret_actual
            sheet['H{}'.format(k)] = energy_ret_actual
            sheet['O{}'.format(k)] = Lat_1ms
            sheet['P{}'.format(k)] = energy_1ms
            sheet['M{}'.format(k)] = exh_lat
            sheet['N{}'.format(k)] = exh_energy
            sheet['R{}'.format(k)] = (Lat_1ms - Lat_pred)/Lat_1ms
            sheet['T{}'.format(k)] = (Lat_1ms - Lat_actual)/Lat_1ms
            sheet['V{}'.format(k)] = (Lat_1ms - exh_lat)/Lat_1ms
            sheet['S{}'.format(k)] = (energy_1ms - energy_pred)/energy_1ms
            sheet['U{}'.format(k)] = (energy_1ms - energy_actual)/energy_1ms
            sheet['W{}'.format(k)] = (energy_1ms - exh_energy)/energy_1ms
            
            k+=1
            #print(f[index],i, j, Lat_pred, Lat_actual)
        
        book.save(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\RF_DT\dcache\Round_{}\dcache_{}_{}.xlsx'.format(number, name, number))
    
