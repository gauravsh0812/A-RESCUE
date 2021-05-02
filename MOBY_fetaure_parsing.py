'''
To parse features for MoBY benchmark suit applications 
and saving it to excel file.
'''

import os
from openpyxl import Workbook

parameters = ['system.mem_ctrls.bytes_read::total', 'system.mem_ctrls.bytes_inst_read::total', 'system.mem_ctrls.num_reads::total', 'system.mem_ctrls.bw_read::total', 'system.mem_ctrls.bw_inst_read::total', 'system.mem_ctrls.bw_total::total', 'system.mem_ctrls.readReqs ','system.mem_ctrls.writeReqs','system.mem_ctrls.bytesReadDRAM','system.mem_ctrls.bytesWritten','system.mem_ctrls.totGap','system.mem_ctrls.totQLat', 'system.mem_ctrls.bytesPerActivate::total','system.mem_ctrls.totMemAccLat','system.mem_ctrls.totBusLat','system.mem_ctrls.avgQLat','system.mem_ctrls.avgBusLat','system.mem_ctrls.avgMemAccLat','system.mem_ctrls_0.totalEnergy', 'system.mem_ctrls_1.totalEnergy','system.cpu.dtb.flush_entries', 'system.cpu.itb.flush_entries', 'system.cpu.numCycles', 'system.cpu.num_int_register_reads','system.cpu.num_int_register_writes', 'system.cpu.num_fp_register_writes', 'system.cpu.num_fp_register_reads', 'system.cpu.num_cc_register_writes', 'system.cpu.num_cc_register_reads', 'system.cpu.dcache.tags.occ_percent::total','system.cpu.dcache.ReadReq_hits::total','system.cpu.dcache.WriteReq_hits::total','system.cpu.dcache.demand_hits::total','system.cpu.dcache.overall_hits::total','system.cpu.dcache.ReadReq_misses::total','system.cpu.dcache.WriteReq_misses::total','system.cpu.dcache.demand_misses::total','system.cpu.dcache.overall_misses::total','system.cpu.dcache.ReadReq_miss_latency::total','system.cpu.dcache.WriteReq_miss_latency::total','system.cpu.dcache.demand_miss_latency::total','system.cpu.dcache.overall_miss_latency::total','system.cpu.dcache.ReadReq_accesses::total','system.cpu.dcache.WriteReq_accesses::total','system.cpu.dcache.demand_accesses::total','system.cpu.dcache.overall_accesses::total','system.cpu.dcache.ReadReq_miss_rate::total','system.cpu.dcache.WriteReq_miss_rate::total','system.cpu.dcache.demand_miss_rate::total','system.cpu.dcache.overall_miss_rate::total','system.cpu.dcache.WriteReq_avg_miss_latency::total','system.cpu.dcache.demand_avg_miss_latency::total','system.cpu.dcache.overall_avg_miss_latency::total','system.cpu.dcache.ReadReq_mshr_misses::total','system.cpu.dcache.WriteReq_mshr_misses::total','system.cpu.dcache.demand_mshr_misses::total','system.cpu.dcache.overall_mshr_misses::total','system.cpu.dcache.ReadReq_mshr_miss_latency::total','system.cpu.dcache.WriteReq_mshr_miss_latency::total','system.cpu.dcache.demand_mshr_miss_latency::total','system.cpu.dcache.overall_mshr_miss_latency::total','system.cpu.dcache.ReadReq_mshr_miss_rate::total','system.cpu.dcache.WriteReq_mshr_miss_rate::total','system.cpu.dcache.demand_mshr_miss_rate::total','system.cpu.dcache.overall_mshr_miss_rate::total','system.cpu.dcache.ReadReq_avg_mshr_miss_latency::total','system.cpu.dcache.WriteReq_avg_mshr_miss_latency::total','system.cpu.dcache.demand_avg_mshr_miss_latency::total','system.cpu.dcache.overall_avg_mshr_miss_latency::total','system.cpu.icache.ReadReq_hits::total','system.cpu.icache.demand_hits::total','system.cpu.icache.overall_hits::total','system.cpu.icache.ReadReq_misses::total','system.cpu.icache.demand_misses::total','system.cpu.icache.overall_misses::total','system.cpu.icache.ReadReq_miss_latency::total','system.cpu.icache.demand_miss_latency::total','system.cpu.icache.overall_miss_latency::total','system.cpu.icache.ReadReq_accesses::total','system.cpu.icache.demand_accesses::total','system.cpu.icache.overall_accesses::total','system.cpu.icache.ReadReq_miss_rate::total','system.cpu.icache.demand_miss_rate::total','system.cpu.icache.overall_miss_rate::total','system.cpu.icache.ReadReq_avg_miss_latency::total','system.cpu.icache.demand_avg_miss_latency::total','system.cpu.icache.overall_avg_miss_latency::total','system.cpu.icache.ReadReq_mshr_misses::total','system.cpu.icache.demand_mshr_misses::total','system.cpu.icache.overall_mshr_misses::total','system.cpu.icache.ReadReq_mshr_miss_latency::total','system.cpu.icache.demand_mshr_miss_latency::total','system.cpu.icache.overall_mshr_miss_latency::total','system.cpu.icache.ReadReq_mshr_miss_rate::total','system.cpu.icache.demand_mshr_miss_rate::total','system.cpu.icache.overall_mshr_miss_rate::total','system.cpu.icache.ReadReq_avg_mshr_miss_latency::total','system.cpu.icache.demand_avg_mshr_miss_latency::total','system.cpu.icache.overall_avg_mshr_miss_latency::total','system.membus.pkt_count::total','system.membus.pkt_size::total','system.switch_cpus.numCycles','system.switch_cpus.committedInsts','system.switch_cpus.committedOps','system.switch_cpus.num_int_alu_accesses','system.switch_cpus.num_fp_alu_accesses','system.switch_cpus.num_func_calls','system.switch_cpus.num_int_register_reads','system.switch_cpus.num_int_register_writes','system.switch_cpus.num_fp_register_reads','system.switch_cpus.num_fp_register_writes','system.switch_cpus.num_cc_register_reads','system.switch_cpus.num_cc_register_writes','system.switch_cpus.op_class::total', 'system.cpu.dcache.overall_miss_rate::total', 'system.cpu.icache.overall_miss_rate::total']
application = ['360buy', 'adobe', 'baidumap', 'frozenbubble', 'k9mail', 'kingsoftoffice', 'mxplayer', 'netease', 'sinaweibo', 'ttpod']
cores = ['1-core', '4-core']
ret_ticks = ['1M', '1G', '1T', '10G','10M', '100G', '100M']
ret_times = ['1us', '1ms', '1s', '10ms', '10us', '100ms', '100us']

def main():
    
    global parameters, applications, cires, ret_ticks, ret_times
    para = []
    value = []
    for i in parameters:
        i=0
    book = Workbook()
    sheet = book.active
    k=2

    for p in range(len(parameters)):
        sheet.cell(row=1, column=p+5).value = parameters[p]

    for core in cores:
        for app in application:
            dir = (r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\MOBY\stats\stats\{}\{}.rcs'.format(core, app))
            os.chdir(dir)
            for r, rr in zip(ret_ticks, ret_times):

                file = '{}.rcs_{}.txt'.format(app,r) 

                print([core, app, file])                
                with open(file,'r') as fp:
                    for line in fp:
                        for j in range(len(parameters)):

                            if parameters[j] in line:
                                end_index = line.find('#')
                                s = line[0:end_index-1]
                                s = s.strip()
                                s = s.split()

                                sheet.cell(row=k, column=1).value = app
                                sheet.cell(row=k, column=2).value = core
                                sheet.cell(row=k, column=3).value = r
                                sheet.cell(row=k, column=4).value = rr
                                sheet.cell(row=k, column=j+5).value = s[1]
                k+=1

    book.save('MOBY_d+i.xlsx')

if __name__=='__main__':
    main()
                    
                                          
