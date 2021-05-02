import pandas as pd
from openpyxl import Workbook

def main()
        book = Workbook()
        sheet = book.active

        arr = ['10us', '26us', '50us', '75us', '1ms']

        k = 1

        df = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\XLSX files\zeusmp_SPEC_data.xlsx')
        df = df.iloc[:,[0,1,3,10,11]]

        for j in range(1,20):            #phase no.
                for i in arr:

                    df_row = df[(df['phase no.'] == j) & (df['ret_time_sec'] == i)]
                    df_row = df_row.iloc[:, [-1,-2]]
                    nor_lat_value = df_row['lat'].values
                    nor_energy_value = df_row['energy'].values

                    sheet.cell(row = k, column = 1).value = j
                    sheet.cell(row = k, column = 2).value = i
                    sheet.cell(row = k, column = 3).value = nor_lat_value[0]
                    sheet.cell(row = k, column = 4).value = nor_energy_value[0]
                    k += 1

        book.save('comaprison.xlsx')

if __name__=='__main__':
        main()
