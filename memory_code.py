# -*- coding: utf-8 -*-
"""
Created on Wed Feb  5 17:44:07 2020

@author: gaura
"""
import pandas as pd
import numpy as np 

df = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\old\DATA_Classifier.xlsx', 'data_75us')
x = df.iloc[:, 3:119]
y = df.iloc[:, -1]
x = pd.DataFrame(x)
x = x.fillna(0)
col = [39,58,40,41,60,19,10,38, 5, 57, 13, 75,76, 82, 71, 83, 43, 33, 30, 35, 31, 64, 109, 92, 68, 52]

from sklearn.preprocessing import StandardScaler
sc_x = StandardScaler()
x = sc_x.fit_transform(x)

from sklearn.model_selection import train_test_split
x_train, x_test, y_train, y_test = train_test_split(x, y, test_size = 0.3, random_state = 0)
x_train = pd.DataFrame(x_train)
x_test = pd.DataFrame(x_test)

x_feat_train = x_train.iloc[:,col]
x_feat_test = x_test.iloc[:,col]

from sklearn.tree import DecisionTreeClassifier
DT = DecisionTreeClassifier( criterion = 'entropy', random_state = 0)
DT.fit(x_feat_train, y_train)

'''import pickle
with open('DT.pickle', 'wb') as dt:
    pickle.dump(DT, dt)
'''

y_pred = DT.predict(x_feat_test)

from sklearn.metrics import confusion_matrix, accuracy_score
cm = confusion_matrix(y_test, y_pred)
accuracy_SVC = accuracy_score(y_test, y_pred)
accuracy_SVC

y_pred_array = np.array(y_pred)
y_pred_array

y_test_array = np.array(y_test)
y_test_array

df_alldata = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\old\DATA_FOR_CLASSIFIER.xlsx', 'data')

index = y_test.index
df_array = []
df_phase = []
df_names = df.iloc[:, 0]
df_phases = df.iloc[:, 2]
for i in index:
    df_array.append(df_names[index])
    df_phase.append(df_phases[index])
    
from openpyxl import Workbook

book = Workbook()
sheet= book.active

df_array = pd.DataFrame(df_array)
df_phase = pd.DataFrame(df_phase)
f = df_array.T.iloc[:,0]
f_phs = df_phase.T.iloc[:,0]
f.columns = ['index', 'app_name']
f_phs.columns = ['index','phase']

index


k = 2

sheet.cell(row = 1, column = 1).value = 'application'
sheet.cell(row = 1, column = 2).value = 'phase'
sheet.cell(row = 1, column = 3).value = 'Predicted_class'
sheet.cell(row = 1, column = 4).value = 'Actual_class'
sheet.cell(row = 1, column = 5).value = 'PREDICTED_LAT_RET'
sheet.cell(row = 1, column = 6).value = 'ACTUAL_LAT_RET'
sheet.cell(row = 1, column = 7).value = 'PRED_ENERGY_RET'
sheet.cell(row = 1, column = 8).value = 'ACTUAL_ENERGY_RET'
sheet.cell(row = 1, column = 9).value = 'Latency_predicted'
sheet.cell(row = 1, column = 10).value = 'Latency_actual'
sheet.cell(row = 1, column = 11).value = 'Energy_predicted'
sheet.cell(row = 1, column = 12).value = 'Energy_actual'

for index, i, j in zip(index, y_pred_array, y_test_array):
    
    sheet['A{}'.format(k)] = f[index]
    sheet['B{}'.format(k)] = f_phs[index]
    sheet['C{}'.format(k)] = i
    sheet['D{}'.format(k)] = j
    
    
    if i == 'D':
        pair = ['10us', '26us']
        
    elif i == 'E':
        pair = ['75us', '1ms']
    
    elif i == 'F':
        pair = ['50us', '75us']
    
    elif i == 'G':
        pair = ['50us', '1ms']
    
    elif i == 'H':
        pair = ['10us', '50us']
    
    elif i == 'I':
        pair = ['26us', '75us']
        
    elif i == 'J':
        pair = ['10us', '75us']    
    
    if j == 'D':
        pair_2 = ['10us', '26us']
        
    elif j == 'E':
        pair_2 = ['75us', '1ms']
    
    elif j == 'F':
        pair_2 = ['50us', '75us']
    
    elif j == 'G':
        pair_2 = ['50us', '1ms']
    
    elif j == 'H':
        pair_2 = ['10us', '50us']
    
    elif j == 'I':
        pair_2 = ['26us', '75us']
        
    elif j == 'J':
        pair_2 = ['10us', '75us']    
        
    lat_ret_pred = pair[1]
    energy_ret_pred = pair[0]    
    lat_ret_actual = pair_2[1]
    energy_ret_actual = pair_2[0]   
    
    
    df_1 = pd.read_excel(r'C:\Users\gaura\OneDrive\Desktop\Simpoint_data_gem5\user_experience_project\SPEC\d_cache\old\DATA_Classifier.xlsx','lat_energy_calc')  
    df_1 = df_1.loc[:,['application','phase no.','ret_time','Latency','ENERGY']]
    
    df_row_lat_pred = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == lat_ret_pred)]
    df_row_energy_pred = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == energy_ret_pred)]
    df_row_lat_actual = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == lat_ret_actual )]
    df_row_energy_actual = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == energy_ret_actual)]
    
    # for base = 1ms
    
    df_row_lat_1ms = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == '1ms')]
    df_row_energy_1ms = df_1.loc[(df_1['application'] == f[index]) & (df_1['phase no.'] == f_phs[index]) & (df_1['ret_time'] == '1ms')]

    Lat_pred = df_row_lat_pred.iloc[:,3]
    Lat_pred = np.array(Lat_pred)
    Lat_pred = Lat_pred[0]
    
    energy_pred = df_row_energy_pred.iloc[:,4]
    energy_pred = np.array(energy_pred)
    energy_pred = energy_pred[0]
    
    Lat_actual = df_row_lat_actual.iloc[:,3]
    Lat_actual = np.array(Lat_actual)
    Lat_actual = Lat_actual[0]
    
    energy_actual = df_row_energy_actual.iloc[:,4]
    energy_actual = np.array(energy_actual)
    energy_actual = energy_actual[0]
    
    Lat_1ms = df_row_lat_1ms.iloc[:,3]
    Lat_1ms = np.array(Lat_1ms)
    Lat_1ms = Lat_1ms[0]
    
    energy_1ms = df_row_energy_1ms.iloc[:,4]
    energy_1ms = np.array(energy_1ms)
    energy_1ms = energy_1ms[0]
    
    sheet['E{}'.format(k)] = lat_ret_pred
    sheet['F{}'.format(k)] = energy_ret_pred
    sheet['I{}'.format(k)] = Lat_pred
    sheet['J{}'.format(k)] = Lat_actual
    sheet['K{}'.format(k)] = energy_pred
    sheet['L{}'.format(k)] = energy_actual
    sheet['G{}'.format(k)] = lat_ret_actual
    sheet['H{}'.format(k)] = energy_ret_actual
    sheet['O{}'.format(k)] = Lat_1ms
    sheet['P{}'.format(k)] = energy_1ms
    
    k+=1
    print(f[index],i, j, Lat_pred, Lat_actual)

book.save('check.xlsx')

'''
from sklearn.model_selection import GridSearchCV

parameters = [
              {'criterion' : ['gini'], 'random_state' : [0] }, 
              {'criterion': ['entropy'], 'random_state' : [0]}]
              

GS_cv = GridSearchCV(estimator = DT, param_grid = parameters, scoring = 'accuracy', cv = 10, n_jobs = -1 )
GS_cv.fit(x_feat_train, y_train)
accuracy = GS_cv.best_score_
GS_cv.best_params_


from memory_profiler import profile

def ml_funct():
    forest = DecisionTreeClassifier( criterion = 'entropy', random_state = 0)
    #gauss = PassiveAggressiveClassifier(random_state=0)
    forest.fit(x_feat_train, y_train)
    
    #a = time.time()
    forest.predict(x_feat_test)
    #b=time.time()
   
    #print(b-a)
    #return mean

if __name__ == '__main__':
    ml_funct()
'''